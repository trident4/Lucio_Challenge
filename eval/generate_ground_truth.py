#!/usr/bin/env python3
"""Excel → ground_truth.json converter.

Reads an Excel file with Question/Answer columns and uses an LLM to generate
structured assertions (contains / contains_any) for automated evaluation.

Usage:
    python eval/generate_ground_truth.py /path/to/questions.xlsx
    python eval/generate_ground_truth.py /path/to/questions.xlsx --corpus "https://..." --dry-run
"""

import argparse
import asyncio
import json
import re
import shutil
import sys
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from openai import AsyncOpenAI

EVAL_DIR = Path(__file__).parent
BACKEND_ENV = EVAL_DIR.parent / "backend" / ".env"
DEFAULT_OUTPUT = EVAL_DIR / "ground_truth.json"

SYSTEM_PROMPT = """\
You extract key verifiable facts from reference answers and return structured
assertions as a JSON array. The assertions are used for automated eval —
they check if an LLM's answer contains the critical facts.

Assertion types (ONLY these two):
- {"type": "contains", "value": "exact string", "label": "description"}
- {"type": "contains_any", "values": ["variant1", "variant2"], "label": "description"}

Rules:
- Use "contains" for exact strings (names, specific numbers, single key terms)
- Use "contains_any" when the fact could be phrased multiple ways
  (e.g., number formatting: "4,586" vs "4586", or "42.3" vs "$42.3B")
- Extract numbers, proper nouns, legal terms, key phrases
- Be thorough: if 9 justices are listed, create 9 separate contains assertions
- For anti-hallucination (answer says info is not available/not in documents):
  generate ONE assertion: {"type": "contains_any",
  "values": ["not available", "not provided", "not exist"],
  "label": "Correctly states information is missing"}
- Return ONLY the JSON array, no markdown fencing"""

FEW_SHOT_EXAMPLES = [
    {
        "role": "user",
        "content": (
            "Question: What are the revenue figures for Meta for Q1, Q2 and Q3?\n"
            "Answer: Meta's revenue was $42.3B in Q1 2025, $39.1B in Q2 2024, "
            "and $40.6B in Q3 2024."
        ),
    },
    {
        "role": "assistant",
        "content": json.dumps(
            [
                {"type": "contains", "value": "42.3", "label": "Q1 2025 revenue $42.3B"},
                {"type": "contains", "value": "39.1", "label": "Q2 2024 revenue $39.1B"},
                {"type": "contains", "value": "40.6", "label": "Q3 2024 revenue $40.6B"},
            ]
        ),
    },
    {
        "role": "user",
        "content": (
            "Question: What was the bench in the Eastman Kodak Case?\n"
            "Answer: The bench included Justices Blackmun, Scalia, O'Connor, "
            "Thomas, Rehnquist, Kennedy, Souter, Stevens, and White."
        ),
    },
    {
        "role": "assistant",
        "content": json.dumps(
            [
                {"type": "contains", "value": "Blackmun", "label": "Justice Blackmun"},
                {"type": "contains", "value": "Scalia", "label": "Justice Scalia"},
                {"type": "contains", "value": "O'Connor", "label": "Justice O'Connor"},
                {"type": "contains", "value": "Thomas", "label": "Justice Thomas"},
                {"type": "contains", "value": "Rehnquist", "label": "Justice Rehnquist"},
                {"type": "contains", "value": "Kennedy", "label": "Justice Kennedy"},
                {"type": "contains", "value": "Souter", "label": "Justice Souter"},
                {"type": "contains", "value": "Stevens", "label": "Justice Stevens"},
                {"type": "contains", "value": "White", "label": "Justice White"},
            ]
        ),
    },
    {
        "role": "user",
        "content": (
            "Question: What was the gross margin for Apple Inc. in Q1 2025?\n"
            "Answer: The information about Apple's gross margin is not available "
            "in the provided documents."
        ),
    },
    {
        "role": "assistant",
        "content": json.dumps(
            [
                {
                    "type": "contains_any",
                    "values": ["not available", "not provided", "not exist"],
                    "label": "Correctly states information is missing",
                }
            ]
        ),
    },
]


# ── Excel Reading ──────────────────────────────────────────────────────────────

QUESTION_VARIANTS = {"question", "questions", "q"}
ANSWER_VARIANTS = {"answer", "answers", "a", "reference answer", "expected answer"}
DOCS_VARIANTS = {"documents referred", "documents", "docs", "docs referred", "sources"}


def _match_column(header: str, variants: set[str]) -> bool:
    return header.strip().lower() in variants


def read_excel(path: str) -> list[dict]:
    """Read Excel file and return list of {question, answer, documents} dicts."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    # Find header row and column indices
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map = {}
    for idx, h in enumerate(headers):
        if h is None:
            continue
        h_str = str(h)
        if _match_column(h_str, QUESTION_VARIANTS):
            col_map["question"] = idx
        elif _match_column(h_str, ANSWER_VARIANTS):
            col_map["answer"] = idx
        elif _match_column(h_str, DOCS_VARIANTS):
            col_map["documents"] = idx

    if "question" not in col_map:
        print("❌ Could not find a 'Question' column in the Excel file.")
        print(f"   Found headers: {headers}")
        sys.exit(1)
    if "answer" not in col_map:
        print("❌ Could not find an 'Answer' column in the Excel file.")
        print(f"   Found headers: {headers}")
        sys.exit(1)

    rows = []
    max_col = max(col_map.values())
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= max_col:
            continue
        q = row[col_map["question"]]
        a = row[col_map["answer"]]
        if not q or not a:
            continue
        entry = {"question": str(q).strip(), "answer": str(a).strip()}
        if "documents" in col_map and len(row) > col_map["documents"] and row[col_map["documents"]]:
            entry["documents"] = str(row[col_map["documents"]]).strip()
        rows.append(entry)

    wb.close()
    return rows


# ── LLM Assertion Generation ──────────────────────────────────────────────────


def _get_client() -> AsyncOpenAI:
    load_dotenv(BACKEND_ENV)
    import os

    return AsyncOpenAI(
        base_url=os.environ["OPENROUTER_BASE_URL"],
        api_key=os.environ["OPENROUTER_API_KEY"],
    )


def _parse_json_array(text: str) -> list[dict] | None:
    """Try to parse a JSON array from LLM response, stripping markdown fences."""
    text = text.strip()
    # Strip markdown code fences if present
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text)
    try:
        result = json.loads(text)
        if isinstance(result, list):
            return result
    except json.JSONDecodeError:
        pass
    return None


async def generate_assertions(
    client: AsyncOpenAI, question: str, answer: str, model: str
) -> list[dict]:
    """Call LLM to generate assertions for a question/answer pair."""
    user_msg = f"Question: {question}\nAnswer: {answer}"
    messages = [
        {"role": "system", "content": SYSTEM_PROMPT},
        *FEW_SHOT_EXAMPLES,
        {"role": "user", "content": user_msg},
    ]

    for attempt in range(2):
        try:
            resp = await client.chat.completions.create(
                model=model,
                messages=messages,
                temperature=0.0,
                max_tokens=1500,
            )
            content = resp.choices[0].message.content
            parsed = _parse_json_array(content)
            if parsed is not None:
                return parsed
            if attempt == 0:
                continue  # retry once
        except Exception as e:
            if attempt == 0:
                continue
            print(f"  ⚠ LLM error for question: {e}")

    # Fallback: simple contains on the full answer
    print(f"  ⚠ Falling back to simple assertion for: {question[:60]}...")
    return [{"type": "contains", "value": answer[:100], "label": "Full answer match (fallback)"}]


async def generate_all_assertions(
    rows: list[dict], model: str
) -> list[list[dict]]:
    """Fire all LLM calls concurrently and return assertions per question."""
    client = _get_client()
    tasks = [
        generate_assertions(client, row["question"], row["answer"], model)
        for row in rows
    ]
    return await asyncio.gather(*tasks)


# ── Assembly & Output ──────────────────────────────────────────────────────────


def build_ground_truth(
    rows: list[dict], all_assertions: list[list[dict]], corpus_url: str
) -> dict:
    questions = []
    for i, (row, assertions) in enumerate(zip(rows, all_assertions), start=1):
        questions.append(
            {
                "id": f"q{i}",
                "text": row["question"],
                "assertions": assertions,
            }
        )
    return {"corpus_url": corpus_url, "questions": questions}


def print_summary(gt: dict) -> None:
    questions = gt["questions"]
    total_assertions = sum(len(q["assertions"]) for q in questions)
    print(f"\n✅ Generated ground_truth.json with {len(questions)} questions, "
          f"{total_assertions} assertions\n")

    print(f"{'Q':<5}| {'Assertions':>10} | Sample")
    print("-" * 70)
    for q in questions:
        a = q["assertions"]
        if not a:
            sample = "(none)"
        elif a[0]["type"] == "contains":
            sample = f'contains "{a[0]["value"]}" ({a[0]["label"]})'
        else:
            vals = a[0].get("values", [])
            sample = f'contains_any {json.dumps(vals)} ({a[0]["label"]})'
        print(f"{q['id']:<5}| {len(a):>10} | {sample}")
    print()


# ── CLI ────────────────────────────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(
        description="Convert Excel questions+answers to ground_truth.json with LLM-generated assertions."
    )
    parser.add_argument("excel_path", help="Path to the Excel file")
    parser.add_argument(
        "--corpus",
        default=None,
        help="Corpus URL/path (default: read from existing ground_truth.json)",
    )
    parser.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT),
        help=f"Output file path (default: {DEFAULT_OUTPUT})",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print generated assertions without writing file",
    )
    parser.add_argument(
        "--model",
        default="openai/gpt-4o-mini",
        help="LLM model to use (default: openai/gpt-4o-mini)",
    )
    args = parser.parse_args()

    # Resolve corpus URL
    corpus_url = args.corpus
    if corpus_url is None:
        if DEFAULT_OUTPUT.exists():
            with open(DEFAULT_OUTPUT) as f:
                corpus_url = json.load(f).get("corpus_url", "")
            print(f"📦 Using corpus URL from existing ground_truth.json: {corpus_url}")
        else:
            corpus_url = ""
            print("⚠ No corpus URL provided and no existing ground_truth.json found.")

    # Read Excel
    print(f"📖 Reading {args.excel_path}...")
    rows = read_excel(args.excel_path)
    if not rows:
        print("❌ No valid questions found in the Excel file.")
        sys.exit(1)
    print(f"   Found {len(rows)} questions\n")

    # Generate assertions
    print(f"🤖 Generating assertions via {args.model} ({len(rows)} concurrent calls)...")
    all_assertions = asyncio.run(generate_all_assertions(rows, args.model))

    # Build ground truth
    gt = build_ground_truth(rows, all_assertions, corpus_url)
    print_summary(gt)

    if args.dry_run:
        print("🔍 Dry run — full output:\n")
        print(json.dumps(gt, indent=2))
        return

    # Backup existing file
    output_path = Path(args.output)
    if output_path.exists():
        backup_path = output_path.with_suffix(".backup.json")
        shutil.copy2(output_path, backup_path)
        print(f"💾 Backed up existing {output_path.name} → {backup_path.name}")

    # Write output
    output_path.write_text(json.dumps(gt, indent=2) + "\n")
    print(f"✅ Written to {output_path}")


if __name__ == "__main__":
    main()
